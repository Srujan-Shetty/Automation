import datetime
import openpyxl
from datetime import datetime
from openpyxl.utils import get_column_letter
import os
from UI_Order_Flow_helper import Get_File_Path
#from NoreShortCutKey import TestCase
class Log:
    #------------Function To create Result Folder--------------------------------#
    def CreateResultFolder(self,Name):   #Create Folder to store Result Log file
        now = datetime.now()
        Timestamp = datetime.timestamp(now)
        Folder_Name = Name + "_" + str(Timestamp)
        directory = Folder_Name
        parent_dir = os.getcwd()+"\Result"
        path = os.path.join(parent_dir, directory)
        os.mkdir(path)
        return path
    #---------------Function to create General_UI_LMT_OrderPlacementTesting.txt and input Build number and Case name --------------#
    def CreateLogTextFile(self,Path):                                 #Create a Result Text File and input TestCase name and Test Build Number
        file_name = "General_UI_LMT_OrderPlacementTesting.txt"            #Initialize Name of the test file
        TestInfo_sheet = Get_File_Path("TestData.xlsx", 0)  # Method to Active the TestData.xlsx sheet
        with (open(os.path.join(Path, file_name), 'a') as fp):        #Function to create Text file in mentioned Path
            Current_Time = Log.Get_current_Time(self)                 #Call a Function to Get Current Timings
            fp.write(str(Current_Time))                               #Write a Timing is the File
            for row1 in range(1, 3):                                  #Number of Row
                for column1 in range(1, 3):                           #Number of Column
                    char = get_column_letter(column1)                 #Function to get Index of the Column
                    print(TestInfo_sheet[str(char) + str(row1)].value)
                    str1 = TestInfo_sheet[char + str(row1)].value
                    fp.write(str(str1))                               #Write values of Cell in .txt file
                fp.write("\n")                                        #Print New Line in .txt file
                Current_Time = Log.Get_current_Time(self)
                fp.write(str(Current_Time))
    #------------------------Function to create Result Excel file "TestResults.xlsx"--------------------------------#
    def CreateExcelFile(self,Path):
        os.makedirs(Path, exist_ok=True)                           # Create the folder if it doesn't exist
        wb = openpyxl.Workbook()                                   # Create a new workbook and select the active sheet
        file_path = os.path.join(Path, "TestResults.xlsx")         # Define the full file path, appending the ".xlsx" extension
        wb.save(file_path)                                         # Save the workbook to the specified file path
        print(f"Excel file saved at: {file_path}")

   #-----------------------Function to write log in "General_UI_LMT_OrderPlacementTesting.txt" --------------------#
    def Write_log_ActualResult_LMTOrder(self,Path,TC_NO,row_value,Test_Scenario,Actual_Qty,Expected_Quantity,Actual_Price,Expected_Price,Actual_OpenQty,Expected_OpenQty,Actual_TradeQty,Expected_TradedQty,Actual_AvgPrice,Expected_AvgPrice,Test_Status,Expected_TransType,Actual_TransType,Expected_TriggerPrice,Actual_TriggerPrice,Expected_DiscQty,Actual_DiscQty):
        file_name = "General_UI_LMT_OrderPlacementTesting.txt"  # Initialize Name of the test file
        TestInfo_sheet = Get_File_Path("TestData.xlsx", 1)  # Method to Active the TestData.xlsx sheet
        print(str(TC_NO) + "----" + str(row_value))
        with (open(os.path.join(Path, file_name), 'a') as fp):
            if Test_Scenario == None:  # Test Scenario Column is None ,increment sub TestCase Number and  do not increment TestCase Number
                print("DONT")
            else:
                fp.write("\n")
                fp.write("#========================================================#")
                fp.write("\n")
                Current_Time = Log.Get_current_Time(self)  # Call a Function to Get Current Timings
                fp.write(str(Current_Time))
                fp.write("TestCaseNumber_"+str(TC_NO)+" -->>")
                fp.write(TestInfo_sheet.cell(row=1, column=1).value)
                fp.write("-->>")
                fp.write(str(TestInfo_sheet.cell(row=row_value, column=1).value))
            fp.write("\n")
            Current_Time = Log.Get_current_Time(self)  # Call a Function to Get Current Timings
            fp.write(str(Current_Time))
            fp.write(str(TestInfo_sheet.cell(row=row_value, column=2).value))
            fp.write("Order Info -->> Done -->> TransType|Qty | Price |Triggerprice|DiscQty|OpenQty |Traded Qty |Fill Avg Price")
            fp.write("\n")
            Current_Time = Log.Get_current_Time(self)  # Call a Function to Get Current Timings
            fp.write(str(Current_Time))
            fp.write("Expected Result -->> Done -->> "+str(Expected_TransType)+ " | "+str(Expected_Quantity) + " | "+str(Expected_Price)+ " | "+str(Expected_TriggerPrice)+ " | " +str(Expected_DiscQty)+ " | "+str(Expected_OpenQty)+ " | "+str(Expected_TradedQty)+ " | "+str(Expected_AvgPrice)+ " | ")
            fp.write("\n")
            Current_Time = Log.Get_current_Time(self)  # Call a Function to Get Current Timings
            fp.write(str(Current_Time))
            fp.write("Actual Result   -->> "+ str(Test_Status)+" -->> "+str(Actual_TransType)+ " | "+str(Actual_Qty) + " | "+str(Actual_Price)+ " | "+str(Actual_TriggerPrice)+ " | "+str(Actual_DiscQty)+ " | "+str(Actual_OpenQty)+ " | "+str(Actual_TradeQty)+ " | "+str(Actual_AvgPrice)+ " | ")
            fp.write("\n")
            fp.write("\n")
        Log.Write_Excel_ActualResult_LMTOrder(self,Path, TC_NO, row_value, Test_Scenario, Test_Status)

    # -----------------------Function to write LMT Order case Result in TestData.xlsx --------------------#
    def Write_Excel_ActualResult_LMTOrder(self,Path,TC_NO,row_value,Test_Scenario,Test_Status):                  #Method to update Testcase result status in "TestResults.xlsx"
        TestInfo_sheet = Get_File_Path("TestData.xlsx", 1)
        Result_Excel_Sheet = os.path.join(Path, "TestResults.xlsx")                                              # "Result_Excel_Sheet" all testcases status will be stored
        if Test_Scenario == None:                                                                                # Test Scenario Column is None ,increment sub TestCase Number and  do not increment TestCase Number
            TestCase = TestInfo_sheet.cell(row=row_value, column=2).value                                        # Access the value from the first row, first column (A1)
            Status = str(Test_Status)                                                                            #PASS/ FAIL Status will be stored in Status
            # Open the target workbook (TestResults.xlsx) to paste the value
            target_workbook = openpyxl.load_workbook(Result_Excel_Sheet)
            target_sheet = target_workbook["LMT_ORDER_TC_STATUS"]
            # Specify target row and column to paste the value
            target_row_index = row_value  # Example: row 3
            target_column_index = 1  # Example: column 2 (B)
            # Paste the value into the target cell
            target_sheet.cell(row=target_row_index, column=target_column_index + 2, value=TestCase)              #Paste TestCase in the 2nd column
            target_sheet.cell(row=target_row_index, column=target_column_index + 3, value=Status)                #Paste Status in 3 rd column

        else:
            print("Executed Else")
            # Access the value from the first row, first column (A1)
            TestCaseNo = str(TC_NO)                                                                              #TestCase Number will be stored in TestCaseNo
            TestScenario = TestInfo_sheet.cell(row=row_value, column=1).value                                    #Get TestScenario from TESTDATA.xlsx and stored in TestScenario
            TestCase = TestInfo_sheet.cell(row=row_value, column=2).value                                        #PASS/FAIL status will be stored in STATUS
            Status = str(Test_Status)
            # Open the target workbook (TestResults.xlsx) to paste the value
            target_workbook = openpyxl.load_workbook(Result_Excel_Sheet)
            if "LMT_ORDER_TC_STATUS" not in target_workbook.sheetnames:
                target_sheet = target_workbook.create_sheet(title="LMT_ORDER_TC_STATUS")
                # Create a second sheet (Sheet2)
                # sheet2 = wb.create_sheet(title="LMT_ORDER_TC_STATUS")  # This creates Sheet2 with a specific name
                target_sheet['A1'] = 'TESTCASE_NUMBER'
                target_sheet['B1'] = 'TEST_SCENARIO'
                target_sheet['C1'] = 'TEST_CASE'
                target_sheet['D1'] = 'STATUS'
            else:
                target_sheet = target_workbook["LMT_ORDER_TC_STATUS"]  # Access Sheet2
            #target_sheet = target_workbook.active  # Assuming data should go into the first sheet
            # Specify target row and column to paste the value
            target_row_index = row_value                   # Example: If row value is 1 then in TestResult.xlsx first row will get updated
            target_column_index = 1                        # Example: column 2 (B)

            # Paste the value into the target cell
            target_sheet.cell(row=target_row_index, column=target_column_index, value=TestCaseNo)
            target_sheet.cell(row=target_row_index, column=target_column_index + 1, value=TestScenario)
            target_sheet.cell(row=target_row_index, column=target_column_index + 2, value=TestCase)
            target_sheet.cell(row=target_row_index, column=target_column_index + 3, value=Status)
            print(str(target_sheet.cell(row=target_row_index, column=target_column_index, value=TestCaseNo)))


        target_workbook.save(Result_Excel_Sheet)                                             # Save the target Excel file with the updated value

    # -----------------------Function to write log in "General_UI_SLLMT_OrderPlacementTesting.txt" --------------------#
    def Write_log_ActualResult_SLLMTOrder(self,Path,TC_NO,row_value,Test_Scenario,Actual_Qty,Expected_Quantity,Actual_Price,Expected_Price,Actual_OpenQty,Expected_OpenQty,Actual_TradeQty,Expected_TradedQty,Actual_AvgPrice,Expected_AvgPrice,Test_Status,Expected_TransType,Actual_TransType,Expected_TriggerPrice,Actual_TriggerPrice,Expected_DiscQty,Actual_DiscQty,Expected_OrderType,Actual_OrderType):
        file_name = "General_UI_SLLMT_OrderPlacementTesting.txt"                        # Initialize Name of the test file
        TestInfo_sheet = Get_File_Path("TestData.xlsx", 2)  # Method to Active the TestData.xlsx sheet
        print(str(TC_NO) + "----" + str(row_value))
        with (open(os.path.join(Path, file_name), 'a') as fp):
            if Test_Scenario == None:                                                   # Test Scenario Column is None ,increment sub TestCase Number and  do not increment TestCase Number
                print("DONT")
            else:
                fp.write("\n")
                fp.write("#========================================================#")
                fp.write("\n")
                Current_Time = Log.Get_current_Time(self)                                # Call a Function to Get Current Timings
                fp.write(str(Current_Time))
                fp.write("TestCaseNumber_" + str(TC_NO) + " -->>")
                fp.write(TestInfo_sheet.cell(row=1, column=1).value)
                fp.write("-->>")
                fp.write(str(TestInfo_sheet.cell(row=row_value, column=1).value))
            fp.write("\n")
            Current_Time = Log.Get_current_Time(self)                                     # Call a Function to Get Current Timings
            fp.write(str(Current_Time))
            fp.write(str(TestInfo_sheet.cell(row=row_value, column=2).value))
            fp.write("Order Info -->> Done -->> TransType|OrdType|Qty | Price |Triggerprice|DiscQty|OpenQty |Traded Qty |Fill Avg Price")
            fp.write("\n")
            Current_Time = Log.Get_current_Time(self)                                       # Call a Function to Get Current Timings
            fp.write(str(Current_Time))
            fp.write("Expected Result -->> Done -->> "+str(Expected_TransType)+ " | "+str(Expected_OrderType) +" | "+str(Expected_Quantity) + " | "+str(Expected_Price)+ " | "+str(Expected_TriggerPrice)+ " | " +str(Expected_DiscQty)+" | "+str(Expected_OpenQty)+ " | "+str(Expected_TradedQty)+ " | "+str(Expected_AvgPrice)+ " | ")
            fp.write("\n")
            Current_Time = Log.Get_current_Time(self)                                       # Call a Function to Get Current Timings
            fp.write(str(Current_Time))
            fp.write("Actual Result   -->> "+ str(Test_Status)+" -->> "+str(Actual_TransType)+" | "+str(Actual_OrderType) + " | "+str(Actual_Qty) + " | "+str(Actual_Price)+ " | "+str(Actual_TriggerPrice)+ " | "+str(Actual_DiscQty)+" | "+str(Actual_OpenQty)+ " | "+str(Actual_TradeQty)+ " | "+str(Actual_AvgPrice)+ " | ")
            fp.write("\n")
            fp.write("\n")
        Log.Write_Excel_ActualResult_SLLMTOrder(self, Path, TC_NO, row_value, Test_Scenario, Test_Status)

    # -----------------------Function to write SLLMT Order case Result in TestData.xlsx --------------------#
    def Write_Excel_ActualResult_SLLMTOrder(self,Path,TC_NO,row_value,Test_Scenario,Test_Status):
        TestInfo_sheet = Get_File_Path("TestData.xlsx", 2)  # Method to Active the TestData.xlsx sheet
        Result_Excel_Sheet = os.path.join(Path,"TestResults.xlsx")                          # "Result_Excel_Sheet" all testcases status will be stored
        if Test_Scenario == None:                                                           # Test Scenario Column is None ,increment sub TestCase Number and  do not increment TestCase Number
            TestCase = TestInfo_sheet.cell(row=row_value,column=2).value                    # Access the value from the first row, first column (A1)
            Status = str(Test_Status)                                                       # PASS/ FAIL Status will be stored in Status
            target_workbook = openpyxl.load_workbook(Result_Excel_Sheet)                    # Open the target workbook (TestResults.xlsx) to paste the value
            target_sheet = target_workbook["SLLMT_ORDER_TC_STATUS"]
            # Specify target row and column to paste the value
            target_row_index = row_value  # Example: row 3
            target_column_index = 1                                                         # Example: column 2 (B)
            # Paste the value into the target cell
            target_sheet.cell(row=target_row_index, column=target_column_index + 2,value=TestCase)  # Paste TestCase in the 2nd column
            target_sheet.cell(row=target_row_index, column=target_column_index + 3,value=Status)    # Paste Status in 3 rd column
        else:
            print("Executed Else")
            # Access the value from the first row, first column (A1)
            TestCaseNo = str(TC_NO)                                                         # TestCase Number will be stored in TestCaseNo
            TestScenario = TestInfo_sheet.cell(row=row_value,column=1).value                # Get TestScenario from TESTDATA.xlsx and stored in TestScenario
            TestCase = TestInfo_sheet.cell(row=row_value, column=2).value                   # PASS/FAIL status will be stored in STATUS
            Status = str(Test_Status)
            # Open the target workbook (TestResults.xlsx) to paste the value
            target_workbook = openpyxl.load_workbook(Result_Excel_Sheet)
            if "SLLMT_ORDER_TC_STATUS" not in target_workbook.sheetnames:
                target_sheet = target_workbook.create_sheet(title="SLLMT_ORDER_TC_STATUS")
                # Create a second sheet (Sheet2)
                #sheet2 = wb.create_sheet(title="SLLMT_ORDER_TC_STATUS")                    # This creates Sheet2 with a specific name
                target_sheet['A1'] = 'TESTCASE_NUMBER'
                target_sheet['B1'] = 'TEST_SCENARIO'
                target_sheet['C1'] = 'TEST_CASE'
                target_sheet['D1'] = 'STATUS'
            else:
                target_sheet = target_workbook["SLLMT_ORDER_TC_STATUS"]                     # Access Sheet2
            # Specify target row and column to paste the value
            target_row_index = row_value                                                    # Example: row value is 1 ,then values will be pasted in row1
            target_column_index = 1                                                         # Example: column 2 (B)

            # Paste the value into the target cell
            target_sheet.cell(row=target_row_index, column=target_column_index, value=TestCaseNo)
            target_sheet.cell(row=target_row_index, column=target_column_index + 1, value=TestScenario)
            target_sheet.cell(row=target_row_index, column=target_column_index + 2, value=TestCase)
            target_sheet.cell(row=target_row_index, column=target_column_index + 3, value=Status)

        target_workbook.save(Result_Excel_Sheet)                                               # Save the target Excel file with the updated value

    # -----------------------Function to write log in "General_UI_MKT_OrderPlacementTesting.txt" --------------------#
    def Write_log_ActualResult_MKTOrder(self, Path, TC_NO, row_value, Test_Scenario, Actual_Qty, Expected_Quantity,Actual_Price, Expected_Price, Actual_OpenQty, Expected_OpenQty, Actual_TradeQty,Expected_TradedQty, Actual_AvgPrice, Expected_AvgPrice, Test_Status,Expected_TransType, Actual_TransType, Expected_TriggerPrice,Actual_TriggerPrice, Expected_DiscQty, Actual_DiscQty, Expected_OrderType,Actual_OrderType):
        file_name = "General_UI_MKT_OrderPlacementTesting.txt"  # Initialize Name of the test file
        TestInfo_sheet = Get_File_Path("TestData.xlsx", 3)  # Method to Active the TestData.xlsx sheet
        print(str(TC_NO) + "----" + str(row_value))

        # Open the file in append mode
        with open(os.path.join(Path, file_name), 'a') as fp:
            if Test_Scenario == None:  # Test Scenario Column is None ,increment sub TestCase Number and do not increment TestCase Number
                print("DONT")
            else:
                fp.write("\n")
                fp.write("#========================================================#")
                fp.write("\n")
                Current_Time = Log.Get_current_Time(self)  # Call a Function to Get Current Timings
                fp.write(str(Current_Time))
                fp.write("TestCaseNumber_" + str(TC_NO) + " -->>")
                fp.write(TestInfo_sheet.cell(row=1, column=1).value)
                fp.write("-->>")
                fp.write(str(TestInfo_sheet.cell(row=row_value, column=1).value))

            # Write the rest of the log
            fp.write("\n")
            Current_Time = Log.Get_current_Time(self)  # Call a Function to Get Current Timings
            fp.write(str(Current_Time))
            fp.write(str(TestInfo_sheet.cell(row=row_value, column=2).value))
            fp.write(
                "Order Info -->> Done -->> TransType|OrdType|Qty | Price |Triggerprice|DiscQty|OpenQty |Traded Qty |Fill Avg Price")
            fp.write("\n")
            Current_Time = Log.Get_current_Time(self)  # Call a Function to Get Current Timings
            fp.write(str(Current_Time))
            fp.write("Expected Result -->> Done -->> " + str(Expected_TransType) + " | " + str(
                Expected_OrderType) + " | " + str(Expected_Quantity) + " | " + str(Expected_Price) + " | " + str(
                Expected_TriggerPrice) + " | " + str(Expected_DiscQty) + " | " + str(Expected_OpenQty) + " | " + str(
                Expected_TradedQty) + " | " + str(Expected_AvgPrice) + " | ")
            fp.write("\n")
            Current_Time = Log.Get_current_Time(self)  # Call a Function to Get Current Timings
            fp.write(str(Current_Time))
            fp.write("Actual Result   -->> " + str(Test_Status) + " -->> " + str(Actual_TransType) + " | " + str(
                Actual_OrderType) + " | " + str(Actual_Qty) + " | " + str(Actual_Price) + " | " + str(
                Actual_TriggerPrice) + " | " + str(Actual_DiscQty) + " | " + str(Actual_OpenQty) + " | " + str(
                Actual_TradeQty) + " | " + str(Actual_AvgPrice) + " | ")
            fp.write("\n")
            fp.write("\n")
        Log.Write_Excel_ActualResult_MKTOrder(self, Path, TC_NO, row_value, Test_Scenario, Test_Status)
    # -----------------------Function to write MKT Order case Result in TestData.xlsx --------------------#
    def Write_Excel_ActualResult_MKTOrder(self,Path,TC_NO,row_value,Test_Scenario,Test_Status):
        TestInfo_sheet = Get_File_Path("TestData.xlsx", 3)  # Method to Active the TestData.xlsx sheet
        Result_Excel_Sheet = os.path.join(Path,"TestResults.xlsx")                          # "Result_Excel_Sheet" all testcases status will be stored
        if Test_Scenario == None:                                                           # Test Scenario Column is None ,increment sub TestCase Number and  do not increment TestCase Number
            TestCase = TestInfo_sheet.cell(row=row_value,column=2).value                    # Access the value from the first row, first column (A1)
            Status = str(Test_Status)                                                       # PASS/ FAIL Status will be stored in Status
            target_workbook = openpyxl.load_workbook(Result_Excel_Sheet)                    # Open the target workbook (TestResults.xlsx) to paste the value
            target_sheet = target_workbook["MKT_ORDER_TC_STATUS"]
            # Specify target row and column to paste the value
            target_row_index = row_value                                                    # Example: row value is 1 ,then values will be pasted in row1
            target_column_index = 1                                                         # Example: column 2 (B)
            # Paste the value into the target cell
            target_sheet.cell(row=target_row_index, column=target_column_index + 2,value=TestCase)  # Paste TestCase in the 2nd column
            target_sheet.cell(row=target_row_index, column=target_column_index + 3,value=Status)    # Paste Status in 3 rd column
        else:
            print("Executed Else")
            # Access the value from the first row, first column (A1)
            TestCaseNo = str(TC_NO)                                                         # TestCase Number will be stored in TestCaseNo
            TestScenario = TestInfo_sheet.cell(row=row_value,column=1).value                # Get TestScenario from TESTDATA.xlsx and stored in TestScenario
            TestCase = TestInfo_sheet.cell(row=row_value, column=2).value                   # PASS/FAIL status will be stored in STATUS
            Status = str(Test_Status)
            # Open the target workbook (TestResults.xlsx) to paste the value
            target_workbook = openpyxl.load_workbook(Result_Excel_Sheet)
            if "MKT_ORDER_TC_STATUS" not in target_workbook.sheetnames:
                target_sheet = target_workbook.create_sheet(title="MKT_ORDER_TC_STATUS")
                # Create a second sheet (Sheet3)
                #sheet3 = wb.create_sheet(title="SLLMT_ORDER_TC_STATUS")                    # This creates Sheet2 with a specific name
                target_sheet['A1'] = 'TESTCASE_NUMBER'
                target_sheet['B1'] = 'TEST_SCENARIO'
                target_sheet['C1'] = 'TEST_CASE'
                target_sheet['D1'] = 'STATUS'
            else:
                target_sheet = target_workbook["MKT_ORDER_TC_STATUS"]                     # Access Sheet2
            # Specify target row and column to paste the value
            target_row_index = row_value                                                    # Example: row value is 1 ,then values will be pasted in row1
            target_column_index = 1                                                         # Example: column 2 (B)

            # Paste the value into the target cell
            target_sheet.cell(row=target_row_index, column=target_column_index, value=TestCaseNo)
            target_sheet.cell(row=target_row_index, column=target_column_index + 1, value=TestScenario)
            target_sheet.cell(row=target_row_index, column=target_column_index + 2, value=TestCase)
            target_sheet.cell(row=target_row_index, column=target_column_index + 3, value=Status)

        target_workbook.save(Result_Excel_Sheet)                                               # Save the target Excel file with the updated value

    # -----------------------Function to write log in "General_UI_SLMKT_OrderPlacementTesting.txt" --------------------#
    def Write_log_ActualResult_SLMKTOrder(self,Path,TC_NO,row_value,Test_Scenario,Actual_Qty,Expected_Quantity,Actual_Price,Expected_Price,Actual_OpenQty,Expected_OpenQty,Actual_TradeQty,Expected_TradedQty,Actual_AvgPrice,Expected_AvgPrice,Test_Status,Expected_TransType,Actual_TransType,Expected_TriggerPrice,Actual_TriggerPrice,Expected_DiscQty,Actual_DiscQty,Expected_OrderType,Actual_OrderType):
        file_name = "General_UI_SLMKT_OrderPlacementTesting.txt"  # Initialize Name of the test file
        TestInfo_sheet = Get_File_Path("TestData.xlsx", 4)  # Method to Active the TestData.xlsx sheet
        print(str(TC_NO) + "----" + str(row_value))
        with (open(os.path.join(Path, file_name), 'a') as fp):
            if Test_Scenario == None:  # Test Scenario Column is None ,increment sub TestCase Number and  do not increment TestCase Number
                print("DONT")
            else:
                fp.write("\n")
                fp.write("#========================================================#")
                fp.write("\n")
                Current_Time = Log.Get_current_Time(self)  # Call a Function to Get Current Timings
                fp.write(str(Current_Time))
                fp.write("TestCaseNumber_" + str(TC_NO) + " -->>")
                fp.write(TestInfo_sheet.cell(row=1, column=1).value)
                fp.write("-->>")
                fp.write(str(TestInfo_sheet.cell(row=row_value, column=1).value))
            fp.write("\n")
            Current_Time = Log.Get_current_Time(self)  # Call a Function to Get Current Timings
            fp.write(str(Current_Time))
            fp.write(str(TestInfo_sheet.cell(row=row_value, column=2).value))
            fp.write("Order Info -->> Done -->> TransType|OrdType|Qty | Price |Triggerprice|DiscQty|OpenQty |Traded Qty |Fill Avg Price")
            fp.write("\n")
            Current_Time = Log.Get_current_Time(self)  # Call a Function to Get Current Timings
            fp.write(str(Current_Time))
            fp.write("Expected Result -->> Done -->> "+str(Expected_TransType)+ " | "+str(Expected_OrderType) +" | "+str(Expected_Quantity) + " | "+str(Expected_Price)+ " | "+str(Expected_TriggerPrice)+ " | " +str(Expected_DiscQty)+" | "+str(Expected_OpenQty)+ " | "+str(Expected_TradedQty)+ " | "+str(Expected_AvgPrice)+ " | ")
            fp.write("\n")
            Current_Time = Log.Get_current_Time(self)  # Call a Function to Get Current Timings
            fp.write(str(Current_Time))
            fp.write("Actual Result   -->> "+ str(Test_Status)+" -->> "+str(Actual_TransType)+" | "+str(Actual_OrderType) + " | "+str(Actual_Qty) + " | "+str(Actual_Price)+ " | "+str(Actual_TriggerPrice)+ " | "+str(Actual_DiscQty)+" | "+str(Actual_OpenQty)+ " | "+str(Actual_TradeQty)+ " | "+str(Actual_AvgPrice)+ " | ")
            fp.write("\n")
            fp.write("\n")
        Log.Write_Excel_ActualResult_SLMKTOrder(self, Path, TC_NO, row_value, Test_Scenario, Test_Status)
    # -----------------------Function to write SLMKT Order case Result in TestData.xlsx --------------------#
    def Write_Excel_ActualResult_SLMKTOrder(self,Path,TC_NO,row_value,Test_Scenario,Test_Status):
        TestInfo_sheet = Get_File_Path("TestData.xlsx", 4)  # Method to Active the TestData.xlsx sheet
        Result_Excel_Sheet = os.path.join(Path,"TestResults.xlsx")                          # "Result_Excel_Sheet" all testcases status will be stored
        if Test_Scenario == None:                                                           # Test Scenario Column is None ,increment sub TestCase Number and  do not increment TestCase Number
            TestCase = TestInfo_sheet.cell(row=row_value,column=2).value                    # Access the value from the first row, first column (A1)
            Status = str(Test_Status)                                                       # PASS/ FAIL Status will be stored in Status
            target_workbook = openpyxl.load_workbook(Result_Excel_Sheet)                    # Open the target workbook (TestResults.xlsx) to paste the value
            target_sheet = target_workbook["SLMKT_ORDER_TC_STATUS"]
            # Specify target row and column to paste the value
            target_row_index = row_value                                                    # Example: row value is 1 ,then values will be pasted in row1
            target_column_index = 1                                                         # Example: column 2 (B)
            # Paste the value into the target cell
            target_sheet.cell(row=target_row_index, column=target_column_index + 2,value=TestCase)  # Paste TestCase in the 2nd column
            target_sheet.cell(row=target_row_index, column=target_column_index + 3,value=Status)    # Paste Status in 3 rd column
        else:
            print("Executed Else")
            # Access the value from the first row, first column (A1)
            TestCaseNo = str(TC_NO)                                                         # TestCase Number will be stored in TestCaseNo
            TestScenario = TestInfo_sheet.cell(row=row_value,column=1).value                # Get TestScenario from TESTDATA.xlsx and stored in TestScenario
            TestCase = TestInfo_sheet.cell(row=row_value, column=2).value                   # PASS/FAIL status will be stored in STATUS
            Status = str(Test_Status)
            # Open the target workbook (TestResults.xlsx) to paste the value
            target_workbook = openpyxl.load_workbook(Result_Excel_Sheet)
            if "SLMKT_ORDER_TC_STATUS" not in target_workbook.sheetnames:
                target_sheet = target_workbook.create_sheet(title="SLMKT_ORDER_TC_STATUS")
                # Create a second sheet (Sheet4)
                #sheet3 = wb.create_sheet(title="SLLMT_ORDER_TC_STATUS")                    # This creates Sheet2 with a specific name
                target_sheet['A1'] = 'TESTCASE_NUMBER'
                target_sheet['B1'] = 'TEST_SCENARIO'
                target_sheet['C1'] = 'TEST_CASE'
                target_sheet['D1'] = 'STATUS'
            else:
                target_sheet = target_workbook["SLMKT_ORDER_TC_STATUS"]                     # Access Sheet2
            # Specify target row and column to paste the value
            target_row_index = row_value                                                    # Example: row value is 1 ,then values will be pasted in row1
            target_column_index = 1                                                         # Example: column 2 (B)

            # Paste the value into the target cell
            target_sheet.cell(row=target_row_index, column=target_column_index, value=TestCaseNo)
            target_sheet.cell(row=target_row_index, column=target_column_index + 1, value=TestScenario)
            target_sheet.cell(row=target_row_index, column=target_column_index + 2, value=TestCase)
            target_sheet.cell(row=target_row_index, column=target_column_index + 3, value=Status)

        target_workbook.save(Result_Excel_Sheet)                                               # Save the target Excel file with the updated value

    #----------------------Function to Get Current Time stamp to write a log ------------------------------------------------------------#
    def Get_current_Time(self):                                       #Function to Get Current Timing to print in a Result file
            current_second = datetime.now().second                        #Initialize current sec in a variable
            current_minute = datetime.now().minute                        #Initialize current min in a variable
            current_hour = datetime.now().hour                            #Initialize current hour in a variable
            current_day = datetime.now().day                              #Initialize current Date in a variable
            current_month = datetime.now().month                          #Initialize current Month in a variable
            current_year = datetime.now().year                            #Initialize current Year in a variable
            current_Time = str(current_day) + "/" + str(current_month) + "/" + str(current_year) + "  " + str(current_hour) + ":" + str(current_minute) + ":" + str(current_second) + " "
            return current_Time


