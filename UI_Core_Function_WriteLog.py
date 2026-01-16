import datetime
import openpyxl
from datetime import datetime
from openpyxl.utils import get_column_letter
import os
from UI_Core_Function_HelperFunction import Get_File_Path
#from NoreShortCutKey import TestCase
class Log:
    #-----------------------------------Function To create Result Folder---------------------------------------#
    def CreateResultFolder(self,Name):   #Create Folder to store Result Log file
        now = datetime.now()
        Timestamp = datetime.timestamp(now)
        Folder_Name = Name + "_" + str(Timestamp)
        directory = Folder_Name
        parent_dir = os.getcwd()+"\Result"
        path = os.path.join(parent_dir, directory)
        os.mkdir(path)
        return path
    #-------------------------Function to create .txt file and input Build number and Case name --------------#
    def CreateLogTextFile(self,Path,file_name,Test_Case):                                 #Create a Result Text File and input TestCase name and Test Build Number
        #file_name = "TC_PRICE_CURSOR.txt"                                     #Initialize Name of the test file
        TestInfo_sheet = Get_File_Path("TestData.xlsx", 0)  # Method to Active the TestData.xlsx sheet
        with (open(os.path.join(Path, file_name), 'a') as fp):        #Function to create Text file in mentioned Path
            Current_Time = Log.Get_current_Time(self)                 #Call a Function to Get Current Timings
            fp.write(str(Current_Time))                               #Write a Timing is the File
            for row1 in range(1, 3):                                  #Number of Row
                for column1 in range(1, 3):                           #Number of Column
                    char = get_column_letter(column1)                 #Function to get Index of the Column
                    print(TestInfo_sheet[str(char) + str(row1)].value)
                    str1 = TestInfo_sheet[char + str(row1)].value
                    fp.write(str(str1))                               #Write values of Cell in .txt file
                fp.write("\n")                                        #Print New Line in .txt file
                Current_Time = Log.Get_current_Time(self)
                fp.write(str(Current_Time))
            fp.write("Edit Price Using ->> "+str(Test_Case))
    #------------------------Function to create Result Excel file "TestResults.xlsx"--------------------------------#
    def CreateExcelFile(self,Path):
        os.makedirs(Path, exist_ok=True)                           # Create the folder if it doesn't exist
        wb = openpyxl.Workbook()                                   # Create a new workbook and select the active sheet
        file_path = os.path.join(Path, "TestResults.xlsx")         # Define the full file path, appending the ".xlsx" extension
        wb.save(file_path)                                         # Save the workbook to the specified file path
        print(f"Excel file saved at: {file_path}")

   #----------------------Function to write initial log in "TC_PRICE_CURSR_SELECT_ANDEDIT_BUY and SELLORDENTRY.txt" --------------------#
    def general_log_select_and_edit(self,Path,Entered_price,Test_Case,row_value,Window_Name,TC_NO,DgtsToselect,Result,Input_Price,Position,Expected_Price,Actual_Price):
        if str(Window_Name) == "Buy Order Entry":
            file_name = "TC_UI_CORE_NORMAL_ORDER_ENTRY_INVOKE_CASES_BUYORDENTRY.txt"  # Initialize Name of the test file
            file_path = os.path.join(Path, "TC_UI_CORE_NORMAL_ORDER_ENTRY_INVOKE_CASES_BUYORDENTRY.txt")
            if not os.path.exists(file_path):
                self.CreateLogTextFile(Path, file_name,Test_Case)
            else:
                print(f"Log file already exists: {file_path}")

            with (open(os.path.join(Path, file_name), 'a') as fp):
                fp.write("\n")
                Current_Time = Log.Get_current_Time(self)  # Call a Function to Get Current Timings
                fp.write(str(Current_Time))
                fp.write("***************** - *******************")
                fp.write("\n")
                Current_Time = Log.Get_current_Time(self)  # Call a Function to Get Current Timings
                fp.write(str(Current_Time))
                fp.write("Price Entered - >>"+str(Entered_price))

            Log.Write_SELECT_AND_EDIT_CASES_Excel(self, Path, TC_NO, Test_Case, DgtsToselect, Window_Name, Result,row_value,Input_Price,Position,Expected_Price,Actual_Price)
        elif str(Window_Name) == "Sell Order Entry":
            file_name = "TC_UI_CORE_NORMAL_ORDER_ENTRY_INVOKE_CASES_SELLORDENTRY.txt"  # Initialize Name of the test file
            file_path = os.path.join(Path, "TC_UI_CORE_NORMAL_ORDER_ENTRY_INVOKE_CASES_SELLORDENTRY.txt")
            if not os.path.exists(file_path):
                self.CreateLogTextFile(Path, file_name, Test_Case)
            else:
                print(f"Log file already exists: {file_path}")

            with (open(os.path.join(Path, file_name), 'a') as fp):
                fp.write("\n")
                Current_Time = Log.Get_current_Time(self)  # Call a Function to Get Current Timings
                fp.write(str(Current_Time))
                fp.write("***************** - *******************")
                fp.write("\n")
                Current_Time = Log.Get_current_Time(self)  # Call a Function to Get Current Timings
                fp.write(str(Current_Time))
                fp.write("Price Entered - >>" + str(Entered_price))
            Log.Write_SELECT_AND_EDIT_CASES_Excel(self, Path, TC_NO, Test_Case, DgtsToselect, Window_Name, Result,row_value, Input_Price, Position, Expected_Price, Actual_Price)

    # ----------------------Function to write initial log in "TC_PRICE_CURSR_DELETE_ANDEDIT_BUY and SELLORDENTRY.txt" --------------------#
    def general_log_delete_and_edit(self,Path,Entered_price,Test_Case,row_value,Window_Name,TC_NO,DgtsToselect,Result,Input_Price,Position,Expected_Price,Actual_Price):
        if str(Window_Name) == "Buy Order Entry":
            file_name = "TC_UI_CORE_NORMAL_ORDER_ENTRY_CLOSE_CASES_BUYORDENTRY.txt"  # Initialize Name of the test file
            file_path = os.path.join(Path, "TC_UI_CORE_NORMAL_ORDER_ENTRY_CLOSE_CASES_BUYORDENTRY.txt")
            if not os.path.exists(file_path):
                self.CreateLogTextFile(Path, file_name, Test_Case)
            else:
                print(f"Log file already exists: {file_path}")

            with (open(os.path.join(Path, file_name), 'a') as fp):
                fp.write("\n")
                Current_Time = Log.Get_current_Time(self)  # Call a Function to Get Current Timings
                fp.write(str(Current_Time))
                fp.write("***************** - *******************")
                fp.write("\n")
                Current_Time = Log.Get_current_Time(self)  # Call a Function to Get Current Timings
                fp.write(str(Current_Time))
                fp.write("Price Entered - >>" + str(Entered_price))
            Log.Write_DELETE_AND_EDIT_CASES_Excel(self, Path, TC_NO, Test_Case, DgtsToselect, Window_Name, Result,row_value, Input_Price, Position, Expected_Price, Actual_Price)
        elif str(Window_Name) == "Sell Order Entry":
            file_name = "TC_UI_CORE_NORMAL_ORDER_ENTRY_CLOSE_CASES_SELLORDENTRY.txt"  # Initialize Name of the test file
            file_path = os.path.join(Path, "TC_UI_CORE_NORMAL_ORDER_ENTRY_CLOSE_CASES_SELLORDENTRY.txt")
            if not os.path.exists(file_path):
                self.CreateLogTextFile(Path, file_name, Test_Case)
            else:
                print(f"Log file already exists: {file_path}")

            with (open(os.path.join(Path, file_name), 'a') as fp):
                fp.write("\n")
                Current_Time = Log.Get_current_Time(self)  # Call a Function to Get Current Timings
                fp.write(str(Current_Time))
                fp.write("***************** - *******************")
                fp.write("\n")
                Current_Time = Log.Get_current_Time(self)  # Call a Function to Get Current Timings
                fp.write(str(Current_Time))
                fp.write("Price Entered - >>" + str(Entered_price))
            Log.Write_DELETE_AND_EDIT_CASES_Excel(self, Path, TC_NO, Test_Case, DgtsToselect, Window_Name, Result,row_value, Input_Price, Position, Expected_Price, Actual_Price)

    # ----------------------Function to write initial log in "TC_PRICE_CURSR_BACKSPACE_ANDEDIT_BUY and SELLORDENTRY.txt" --------------------#
    def general_log_backspace_and_edit(self,Path,Entered_price,Test_Case,row_value,Window_Name,TC_NO,DgtsToselect,Result,Input_Price,Position,Expected_Price,Actual_Price):
        if str(Window_Name) == "Buy Order Entry":
            file_name = "TC_UI_CORE_NORMAL_ORDER_ENTRY_MINIMIZE_CASES_BUYORDENTRY.txt"  # Initialize Name of the test file
            file_path = os.path.join(Path, "TC_UI_CORE_NORMAL_ORDER_ENTRY_MINIMIZE_CASES_BUYORDENTRY.txt")
            if not os.path.exists(file_path):
                self.CreateLogTextFile(Path, file_name, Test_Case)
            else:
                print(f"Log file already exists: {file_path}")

            with (open(os.path.join(Path, file_name), 'a') as fp):
                fp.write("\n")
                Current_Time = Log.Get_current_Time(self)  # Call a Function to Get Current Timings
                fp.write(str(Current_Time))
                fp.write("***************** - *******************")
                fp.write("\n")
                Current_Time = Log.Get_current_Time(self)  # Call a Function to Get Current Timings
                fp.write(str(Current_Time))
                fp.write("Price Entered - >>" + str(Entered_price))
            Log.Write_BACKSPACE_AND_EDIT_CASES_Excel(self, Path, TC_NO, Test_Case, DgtsToselect, Window_Name, Result,row_value, Input_Price, Position, Expected_Price, Actual_Price)
        elif str(Window_Name) == "Sell Order Entry":
            file_name = "TC_UI_CORE_NORMAL_ORDER_ENTRY_MINIMIZE_CASES_SELLORDENTRY.txt"  # Initialize Name of the test file
            file_path = os.path.join(Path, "TC_UI_CORE_NORMAL_ORDER_ENTRY_MINIMIZE_CASES_SELLORDENTRY.txt")
            if not os.path.exists(file_path):
                self.CreateLogTextFile(Path, file_name, Test_Case)
            else:
                print(f"Log file already exists: {file_path}")

            with (open(os.path.join(Path, file_name), 'a') as fp):
                fp.write("\n")
                Current_Time = Log.Get_current_Time(self)  # Call a Function to Get Current Timings
                fp.write(str(Current_Time))
                fp.write("***************** - *******************")
                fp.write("\n")
                Current_Time = Log.Get_current_Time(self)  # Call a Function to Get Current Timings
                fp.write(str(Current_Time))
                fp.write("Price Entered - >>" + str(Entered_price))
            Log.Write_BACKSPACE_AND_EDIT_CASES_Excel(self, Path, TC_NO, Test_Case, DgtsToselect, Window_Name, Result,row_value, Input_Price, Position, Expected_Price, Actual_Price)
    # ----------------------Function to write initial log in "TC_PRICE_CURSR_INCREMENT_BUY and SELLORDENTRY.txt" --------------------#
    def general_log_Increment(self,Path,Entered_price,Test_Case,row_value,Window_Name,Exch,TC_NO, DgtsToselect, Result,Expected_Price, Actual_Price):
        if str(Window_Name) == "Buy Order Entry":
            file_name = "TC_UI_CORE_NORMAL_ORDER_ENTRY_MAXIMIZE_CASES_BUYORDENTRY.txt"  # Initialize Name of the test file
            file_path = os.path.join(Path, "TC_UI_CORE_NORMAL_ORDER_ENTRY_MAXIMIZE_CASES_BUYORDENTRY.txt")
            if not os.path.exists(file_path):
                self.CreateLogTextFile(Path, file_name, Test_Case)
            else:
                print(f"Log file already exists: {file_path}")

            with (open(os.path.join(Path, file_name), 'a') as fp):
                fp.write("\n")
                Current_Time = Log.Get_current_Time(self)                    # Call a Function to Get Current Timings
                fp.write(str(Current_Time))
                fp.write("***************** - *******************")
                fp.write("\n")
                Current_Time = Log.Get_current_Time(self)                   # Call a Function to Get Current Timings
                fp.write(str(Current_Time))
                fp.write("Price Entered - >>" + str(Entered_price))
                fp.write("\n")
                Current_Time = Log.Get_current_Time(self)                   # Call a Function to Get Current Timings
                fp.write(str(Current_Time))
                fp.write("Exchange - >>" + str(Exch))
            Log.Write_UP_ARROW_INCREMENT_CASES_Excel(self, Path, TC_NO, Test_Case, DgtsToselect, Window_Name, Result,row_value,Expected_Price, Actual_Price,Exch)
        elif str(Window_Name) == "Sell Order Entry":
            file_name = "TC_UI_CORE_NORMAL_ORDER_ENTRY_MAXIMIZE_CASES_SELLORDENTRY.txt"  # Initialize Name of the test file
            file_path = os.path.join(Path, "TC_UI_CORE_NORMAL_ORDER_ENTRY_MAXIMIZE_CASES_SELLORDENTRY.txt")
            if not os.path.exists(file_path):
                self.CreateLogTextFile(Path, file_name, Test_Case)
            else:
                print(f"Log file already exists: {file_path}")

            with (open(os.path.join(Path, file_name), 'a') as fp):
                fp.write("\n")
                Current_Time = Log.Get_current_Time(self)                   # Call a Function to Get Current Timings
                fp.write(str(Current_Time))
                fp.write("***************** - *******************")
                fp.write("\n")
                Current_Time = Log.Get_current_Time(self)  # Call a Function to Get Current Timings
                fp.write(str(Current_Time))
                fp.write("Price Entered - >>" + str(Entered_price))
                fp.write("\n")
                Current_Time = Log.Get_current_Time(self)                   # Call a Function to Get Current Timings
                fp.write(str(Current_Time))
                fp.write("Exchange - >>" + str(Exch))
            Log.Write_UP_ARROW_INCREMENT_CASES_Excel(self, Path, TC_NO, Test_Case, DgtsToselect, Window_Name, Result,row_value,Expected_Price, Actual_Price,Exch)

    # ----------------------Function to write initial log in "TC_PRICE_CURSR_DECREMENT_BUY and SELLORDENTRY.txt" --------------------#
    def general_log_Decrement(self,Path,Entered_price,Test_Case,row_value,Window_Name,Exch,TC_NO, DgtsToselect, Result,Expected_Price, Actual_Price):
        if str(Window_Name) == "Buy Order Entry":
            file_name = "TC_UI_CORE_NORMAL_ORDER_ENTRY_TAB_FUNCTIONALITY_BUYORDENTRY.txt"  # Initialize Name of the test file
            file_path = os.path.join(Path, "TC_UI_CORE_NORMAL_ORDER_ENTRY_TAB_FUNCTIONALITY_BUYORDENTRY.txt")
            if not os.path.exists(file_path):
                self.CreateLogTextFile(Path, file_name, Test_Case)
            else:
                print(f"Log file already exists: {file_path}")

            with (open(os.path.join(Path, file_name), 'a') as fp):
                fp.write("\n")
                Current_Time = Log.Get_current_Time(self)                    # Call a Function to Get Current Timings
                fp.write(str(Current_Time))
                fp.write("***************** - *******************")
                fp.write("\n")
                Current_Time = Log.Get_current_Time(self)                   # Call a Function to Get Current Timings
                fp.write(str(Current_Time))
                fp.write("Price Entered - >>" + str(Entered_price))
                fp.write("\n")
                Current_Time = Log.Get_current_Time(self)                   # Call a Function to Get Current Timings
                fp.write(str(Current_Time))
                fp.write("Exchange - >>" + str(Exch))
            Log.Write_DOWN_ARROW_DECREMENT_CASES_Excel(self, Path, TC_NO, Test_Case, DgtsToselect, Window_Name, Result,row_value,Expected_Price, Actual_Price,Exch)
        elif str(Window_Name) == "Sell Order Entry":
            file_name = "TC_UI_CORE_NORMAL_ORDER_ENTRY_TAB_FUNCTIONALITY_BUYORDENTRY.txtSELLORDENTRY.txt"  # Initialize Name of the test file
            file_path = os.path.join(Path, "TC_UI_CORE_NORMAL_ORDER_ENTRY_TAB_FUNCTIONALITY_BUYORDENTRY.txt.txt")
            if not os.path.exists(file_path):
                self.CreateLogTextFile(Path, file_name, Test_Case)
            else:
                print(f"Log file already exists: {file_path}")

            with (open(os.path.join(Path, file_name), 'a') as fp):
                fp.write("\n")
                Current_Time = Log.Get_current_Time(self)                   # Call a Function to Get Current Timings
                fp.write(str(Current_Time))
                fp.write("***************** - *******************")
                fp.write("\n")
                Current_Time = Log.Get_current_Time(self)  # Call a Function to Get Current Timings
                fp.write(str(Current_Time))
                fp.write("Price Entered - >>" + str(Entered_price))
                fp.write("\n")
                Current_Time = Log.Get_current_Time(self)                   # Call a Function to Get Current Timings
                fp.write(str(Current_Time))
                fp.write("Exchange - >>" + str(Exch))
            Log.Write_DOWN_ARROW_DECREMENT_CASES_Excel(self,Path, TC_NO, Test_Case, DgtsToselect, Window_Name, Result,row_value, Expected_Price, Actual_Price,Exch)

    # ----------------------Function to write initial log in "TC_PRICE_CURSR_RIGHTARROW_INSERT_BUY and SELLORDENTRY.txt" --------------------#
    def general_log_move_rightarrow_and_Insert(self,Path,Entered_price,Test_Case,row_value,Window_Name,TC_NO,DgtsToselect,Result,Input_Price,Position,Expected_Price,Actual_Price):
        if str(Window_Name) == "Buy Order Entry":
            file_name = "TC_UI_CORE_NORMAL_ORDER-ENTRY_REVERSE_TAB_FUNCTIONALITY_BUYORDENTRY.txt"  # Initialize Name of the test file
            file_path = os.path.join(Path, "TC_UI_CORE_NORMAL_ORDER-ENTRY_REVERSE_TAB_FUNCTIONALITY_BUYORDENTRY.txt")
            if not os.path.exists(file_path):
                self.CreateLogTextFile(Path, file_name, Test_Case)
            else:
                print(f"Log file already exists: {file_path}")

            with (open(os.path.join(Path, file_name), 'a') as fp):
                fp.write("\n")
                Current_Time = Log.Get_current_Time(self)                    # Call a Function to Get Current Timings
                fp.write(str(Current_Time))
                fp.write("***************** - *******************")
                fp.write("\n")
                Current_Time = Log.Get_current_Time(self)                   # Call a Function to Get Current Timings
                fp.write(str(Current_Time))
                fp.write("Price Entered - >>" + str(Entered_price))
                fp.write("\n")
            Log.Write_RIGHTARROW_AND_INSERT_CASES_Excel(self, Path, TC_NO, Test_Case, DgtsToselect, Window_Name, Result,row_value, Input_Price, Position, Expected_Price, Actual_Price)
        elif str(Window_Name) == "Sell Order Entry":
            file_name = "TC_UI_CORE_NORMAL_ORDER-ENTRY_REVERSE_TAB_FUNCTIONALITY_SELLORDENTRY.txt"  # Initialize Name of the test file
            file_path = os.path.join(Path, "TC_UI_CORE_NORMAL_ORDER-ENTRY_REVERSE_TAB_FUNCTIONALITY_SELLORDENTRY.txt")
            if not os.path.exists(file_path):
                self.CreateLogTextFile(Path, file_name, Test_Case)
            else:
                print(f"Log file already exists: {file_path}")

            with (open(os.path.join(Path, file_name), 'a') as fp):
                fp.write("\n")
                Current_Time = Log.Get_current_Time(self)                   # Call a Function to Get Current Timings
                fp.write(str(Current_Time))
                fp.write("***************** - *******************")
                fp.write("\n")
                Current_Time = Log.Get_current_Time(self)  # Call a Function to Get Current Timings
                fp.write(str(Current_Time))
                fp.write("Price Entered - >>" + str(Entered_price))
                fp.write("\n")
            Log.Write_RIGHTARROW_AND_INSERT_CASES_Excel(self, Path, TC_NO, Test_Case, DgtsToselect, Window_Name, Result,row_value, Input_Price, Position, Expected_Price, Actual_Price)
#==================================================================================================================================================================================================#
    # ----------------------Function to write log in "TC_PRICE_CURSR_SELECT_ANDEDIT_BUY and SELLORDENTRY.txt" --------------------#
    def Write_SELECT_AND_EDIT_CASES(self,Path,TC_NO,WINDOW_NAME,TestCase,Input_Price,DgtsToselect,Position,Expected_Price,Actual_Price,Result,row_value):
        if str(WINDOW_NAME) == "Buy Order Entry":
            file_name = "TC_PRICE_CURSR_SELECT_ANDEDIT_BUYORDENTRY.txt"            # Initialize Name of the test file
            TestInfo_sheet = Get_File_Path("TESTCASE.xlsx", 1)  # Method to Active the TestData.xlsx sheet
            #print(str(TC_NO) + "----" + str(row_value))
            with (open(os.path.join(Path, file_name), 'a') as fp):
                fp.write("\n")
                Current_Time = Log.Get_current_Time(self)           # Call a Function to Get Current Timings
                fp.write(str(Current_Time))
                fp.write("Window Name ->> "+ str(WINDOW_NAME))
                fp.write("\n")
                Current_Time = Log.Get_current_Time(self)                  # Call a Function to Get Current Timings
                fp.write(str(Current_Time))
                fp.write("Scenario ->> "+ str(DgtsToselect)+" and Position of digit is "+str(Position))
                fp.write("\n")
                Current_Time = Log.Get_current_Time(self)                 # Call a Function to Get Current Timings
                fp.write(str(Current_Time))
                fp.write("Expected PriceAfterEditing ->> "+ str(Expected_Price))
                fp.write("\n")
                Current_Time = Log.Get_current_Time(self)                 # Call a Function to Get Current Timings
                fp.write(str(Current_Time))
                fp.write("Actual PriceAfterEditing ->> "+ str(Actual_Price))
                fp.write("\n")
                Current_Time = Log.Get_current_Time(self)                  # Call a Function to Get Current Timings
                fp.write(str(Current_Time))
                fp.write("Result ->> " + str(Result))
                fp.write("\n")
                Current_Time = Log.Get_current_Time(self)                  # Call a Function to Get Current Timings
                fp.write(str(Current_Time))
                fp.write("***************** - *******************")
            Log.Write_SELECT_AND_EDIT_CASES_Excel(self, Path, TC_NO, TestCase, DgtsToselect, WINDOW_NAME, Result,row_value,Input_Price,Position,Expected_Price,Actual_Price)
        elif str(WINDOW_NAME) == "Sell Order Entry":
            file_name = "TC_PRICE_CURSR_SELECT_ANDEDIT_SELLORDENTRY.txt"  # Initialize Name of the test file
            TestInfo_sheet = Get_File_Path("TESTCASE.xlsx", 1)  # Method to Active the TestData.xlsx sheet
            # print(str(TC_NO) + "----" + str(row_value))
            with (open(os.path.join(Path, file_name), 'a') as fp):
                fp.write("\n")
                Current_Time = Log.Get_current_Time(self)  # Call a Function to Get Current Timings
                fp.write(str(Current_Time))
                fp.write("Window Name ->> " + str(WINDOW_NAME))
                fp.write("\n")
                Current_Time = Log.Get_current_Time(self)  # Call a Function to Get Current Timings
                fp.write(str(Current_Time))
                fp.write("Scenario ->> " + str(DgtsToselect) + " and Position of digit is " + str(Position))
                fp.write("\n")
                Current_Time = Log.Get_current_Time(self)  # Call a Function to Get Current Timings
                fp.write(str(Current_Time))
                fp.write("Expected PriceAfterEditing ->> " + str(Expected_Price))
                fp.write("\n")
                Current_Time = Log.Get_current_Time(self)  # Call a Function to Get Current Timings
                fp.write(str(Current_Time))
                fp.write("Actual PriceAfterEditing ->> " + str(Actual_Price))
                fp.write("\n")
                Current_Time = Log.Get_current_Time(self)  # Call a Function to Get Current Timings
                fp.write(str(Current_Time))
                fp.write("Result ->> " + str(Result))
                fp.write("\n")
                Current_Time = Log.Get_current_Time(self)  # Call a Function to Get Current Timings
                fp.write(str(Current_Time))
                fp.write("***************** - *******************")
            Log.Write_SELECT_AND_EDIT_CASES_Excel(self, Path, TC_NO, TestCase, DgtsToselect, WINDOW_NAME, Result,row_value,Input_Price,Position,Expected_Price,Actual_Price)

    # ----------------------Function to write log in "TC_PRICE_CURSR_DELETE_ANDEDIT_BUY and SELLORDENTRY.txt" --------------------#
    def Write_DELETE_AND_EDIT_CASES(self,Path,TC_NO,WINDOW_NAME,TestCase,Input_Price,DgtsToselect,Position,Expected_Price,Actual_Price,Result,row_value):
        if str(WINDOW_NAME) == "Buy Order Entry":
            file_name = "TC_PRICE_CURSR_DELETE_ANDEDIT_BUYORDENTRY.txt"  # Initialize Name of the test file
            TestInfo_sheet = Get_File_Path("TESTCASE.xlsx", 1)  # Method to Active the TestData.xlsx sheet
            # print(str(TC_NO) + "----" + str(row_value))
            with (open(os.path.join(Path, file_name), 'a') as fp):
                fp.write("\n")
                Current_Time = Log.Get_current_Time(self)  # Call a Function to Get Current Timings
                fp.write(str(Current_Time))
                fp.write("Window Name ->> " + str(WINDOW_NAME))
                fp.write("\n")
                Current_Time = Log.Get_current_Time(self)  # Call a Function to Get Current Timings
                fp.write(str(Current_Time))
                fp.write("Scenario ->> " + str(DgtsToselect) + " and Position of digit is " + str(Position))
                fp.write("\n")
                Current_Time = Log.Get_current_Time(self)  # Call a Function to Get Current Timings
                fp.write(str(Current_Time))
                fp.write("Expected PriceAfterEditing ->> " + str(Expected_Price))
                fp.write("\n")
                Current_Time = Log.Get_current_Time(self)  # Call a Function to Get Current Timings
                fp.write(str(Current_Time))
                fp.write("Actual PriceAfterEditing ->> " + str(Actual_Price))
                fp.write("\n")
                Current_Time = Log.Get_current_Time(self)  # Call a Function to Get Current Timings
                fp.write(str(Current_Time))
                fp.write("Result ->> " + str(Result))
                fp.write("\n")
                Current_Time = Log.Get_current_Time(self)  # Call a Function to Get Current Timings
                fp.write(str(Current_Time))
                fp.write("***************** - *******************")
            Log.Write_DELETE_AND_EDIT_CASES_Excel(self, Path, TC_NO, TestCase, DgtsToselect, WINDOW_NAME, Result,row_value, Input_Price, Position, Expected_Price, Actual_Price)
        elif str(WINDOW_NAME) == "Sell Order Entry":
            file_name = "TC_PRICE_CURSR_DELETE_ANDEDIT_SELLORDENTRY.txt"  # Initialize Name of the test file
            TestInfo_sheet = Get_File_Path("TESTCASE.xlsx", 1)  # Method to Active the TestData.xlsx sheet
            # print(str(TC_NO) + "----" + str(row_value))
            with (open(os.path.join(Path, file_name), 'a') as fp):
                fp.write("\n")
                Current_Time = Log.Get_current_Time(self)  # Call a Function to Get Current Timings
                fp.write(str(Current_Time))
                fp.write("Window Name ->> " + str(WINDOW_NAME))
                fp.write("\n")
                Current_Time = Log.Get_current_Time(self)  # Call a Function to Get Current Timings
                fp.write(str(Current_Time))
                fp.write("Scenario ->> " + str(DgtsToselect) + " and Position of digit is " + str(Position))
                fp.write("\n")
                Current_Time = Log.Get_current_Time(self)  # Call a Function to Get Current Timings
                fp.write(str(Current_Time))
                fp.write("Expected PriceAfterEditing ->> " + str(Expected_Price))
                fp.write("\n")
                Current_Time = Log.Get_current_Time(self)  # Call a Function to Get Current Timings
                fp.write(str(Current_Time))
                fp.write("Actual PriceAfterEditing ->> " + str(Actual_Price))
                fp.write("\n")
                Current_Time = Log.Get_current_Time(self)  # Call a Function to Get Current Timings
                fp.write(str(Current_Time))
                fp.write("Result ->> " + str(Result))
                fp.write("\n")
                Current_Time = Log.Get_current_Time(self)  # Call a Function to Get Current Timings
                fp.write(str(Current_Time))
                fp.write("***************** - *******************")
            Log.Write_DELETE_AND_EDIT_CASES_Excel(self, Path, TC_NO, TestCase, DgtsToselect, WINDOW_NAME, Result,row_value, Input_Price, Position, Expected_Price, Actual_Price)

    # ----------------------Function to write log in "TC_PRICE_CURSR_BACKSPACE_ANDEDIT_BUY and SELLORDENTRY.txt" --------------------#
    def Write_BACKSPACE_AND_EDIT_CASES(self,Path,TC_NO,WINDOW_NAME,TestCase,Input_Price,DgtsToselect,Position,Expected_Price,Actual_Price,Result,row_value):
        if str(WINDOW_NAME) == "Buy Order Entry":
            file_name = "TC_PRICE_CURSR_BACKSPACE_ANDEDIT_BUYORDENTRY.txt"  # Initialize Name of the test file
            TestInfo_sheet = Get_File_Path("TESTCASE.xlsx", 1)  # Method to Active the TestData.xlsx sheet
            # print(str(TC_NO) + "----" + str(row_value))
            with (open(os.path.join(Path, file_name), 'a') as fp):
                fp.write("\n")
                Current_Time = Log.Get_current_Time(self)  # Call a Function to Get Current Timings
                fp.write(str(Current_Time))
                fp.write("Window Name ->> " + str(WINDOW_NAME))
                fp.write("\n")
                Current_Time = Log.Get_current_Time(self)  # Call a Function to Get Current Timings
                fp.write(str(Current_Time))
                fp.write("Scenario ->> " + str(DgtsToselect) + " and Position of digit is " + str(Position))
                fp.write("\n")
                Current_Time = Log.Get_current_Time(self)  # Call a Function to Get Current Timings
                fp.write(str(Current_Time))
                fp.write("Expected PriceAfterEditing ->> " + str(Expected_Price))
                fp.write("\n")
                Current_Time = Log.Get_current_Time(self)  # Call a Function to Get Current Timings
                fp.write(str(Current_Time))
                fp.write("Actual PriceAfterEditing ->> " + str(Actual_Price))
                fp.write("\n")
                Current_Time = Log.Get_current_Time(self)  # Call a Function to Get Current Timings
                fp.write(str(Current_Time))
                fp.write("Result ->> " + str(Result))
                fp.write("\n")
                Current_Time = Log.Get_current_Time(self)  # Call a Function to Get Current Timings
                fp.write(str(Current_Time))
                fp.write("***************** - *******************")
            Log.Write_BACKSPACE_AND_EDIT_CASES_Excel(self, Path, TC_NO, TestCase, DgtsToselect, WINDOW_NAME, Result,row_value, Input_Price, Position, Expected_Price, Actual_Price)
        elif str(WINDOW_NAME) == "Sell Order Entry":
            file_name = "TC_PRICE_CURSR_BACKSPACE_ANDEDIT_SELLORDENTRY.txt"  # Initialize Name of the test file
            TestInfo_sheet = Get_File_Path("TESTCASE.xlsx", 1)  # Method to Active the TestData.xlsx sheet
            # print(str(TC_NO) + "----" + str(row_value))
            with (open(os.path.join(Path, file_name), 'a') as fp):
                fp.write("\n")
                Current_Time = Log.Get_current_Time(self)  # Call a Function to Get Current Timings
                fp.write(str(Current_Time))
                fp.write("Window Name ->> " + str(WINDOW_NAME))
                fp.write("\n")
                Current_Time = Log.Get_current_Time(self)  # Call a Function to Get Current Timings
                fp.write(str(Current_Time))
                fp.write("Scenario ->> " + str(DgtsToselect) + " and Position of digit is " + str(Position))
                fp.write("\n")
                Current_Time = Log.Get_current_Time(self)  # Call a Function to Get Current Timings
                fp.write(str(Current_Time))
                fp.write("Expected PriceAfterEditing ->> " + str(Expected_Price))
                fp.write("\n")
                Current_Time = Log.Get_current_Time(self)  # Call a Function to Get Current Timings
                fp.write(str(Current_Time))
                fp.write("Actual PriceAfterEditing ->> " + str(Actual_Price))
                fp.write("\n")
                Current_Time = Log.Get_current_Time(self)  # Call a Function to Get Current Timings
                fp.write(str(Current_Time))
                fp.write("Result ->> " + str(Result))
                fp.write("\n")
                Current_Time = Log.Get_current_Time(self)  # Call a Function to Get Current Timings
                fp.write(str(Current_Time))
                fp.write("***************** - *******************")
            Log.Write_BACKSPACE_AND_EDIT_CASES_Excel(self, Path, TC_NO, TestCase, DgtsToselect, WINDOW_NAME, Result,row_value, Input_Price, Position, Expected_Price, Actual_Price)

    # ----------------------Function to write log in "TC_PRICE_CURSR_INCREMENT_BUY and SELLORDENTRY.txt" --------------------#
    def Write_UP_ARROW_INCREMENT_CASES(self,Path, TC_NO,WINDOW_NAME,Test_Case, Exch,row_value, DgtsToselect, Result,Expected_Price, Actual_Price):
        if str(WINDOW_NAME) == "Buy Order Entry":
            file_name = "TC_PRICE_CURSR_INCREMENT_BUYORDENTRY.txt"  # Initialize Name of the test file
            TestInfo_sheet = Get_File_Path("TESTCASE.xlsx", 1)  # Method to Active the TestData.xlsx sheet
            # print(str(TC_NO) + "----" + str(row_value))
            with (open(os.path.join(Path, file_name), 'a') as fp):
                fp.write("\n")
                Current_Time = Log.Get_current_Time(self)  # Call a Function to Get Current Timings
                fp.write(str(Current_Time))
                fp.write("Window Name ->> " + str(WINDOW_NAME))
                fp.write("\n")
                Current_Time = Log.Get_current_Time(self)  # Call a Function to Get Current Timings
                fp.write(str(Current_Time))
                fp.write("Scenario ->> Price " + str(DgtsToselect) + " using Up Arrow ")
                fp.write("\n")
                Current_Time = Log.Get_current_Time(self)  # Call a Function to Get Current Timings
                fp.write(str(Current_Time))
                fp.write("Expected PriceAfterIncrement ->> " + str(Expected_Price))
                fp.write("\n")
                Current_Time = Log.Get_current_Time(self)  # Call a Function to Get Current Timings
                fp.write(str(Current_Time))
                fp.write("Actual PriceAfterIncrement ->> " + str(Actual_Price))
                fp.write("\n")
                Current_Time = Log.Get_current_Time(self)  # Call a Function to Get Current Timings
                fp.write(str(Current_Time))
                fp.write("Result ->> " + str(Result))
                fp.write("\n")
                Current_Time = Log.Get_current_Time(self)  # Call a Function to Get Current Timings
                fp.write(str(Current_Time))
                fp.write("***************** - *******************")
            Log.Write_UP_ARROW_INCREMENT_CASES_Excel(self, Path, TC_NO, Test_Case, DgtsToselect, WINDOW_NAME, Result,row_value, Expected_Price, Actual_Price,Exch)
        elif str(WINDOW_NAME) == "Sell Order Entry":
            file_name = "TC_PRICE_CURSR_INCREMENT_SELLORDENTRY.txt"  # Initialize Name of the test file
            TestInfo_sheet = Get_File_Path("TESTCASE.xlsx", 1)  # Method to Active the TestData.xlsx sheet
            # print(str(TC_NO) + "----" + str(row_value))
            with (open(os.path.join(Path, file_name), 'a') as fp):
                fp.write("\n")
                Current_Time = Log.Get_current_Time(self)  # Call a Function to Get Current Timings
                fp.write(str(Current_Time))
                fp.write("Window Name ->> " + str(WINDOW_NAME))
                fp.write("\n")
                Current_Time = Log.Get_current_Time(self)  # Call a Function to Get Current Timings
                fp.write(str(Current_Time))
                fp.write("Scenario ->> Price " + str(DgtsToselect) + " using Up Arrow ")
                fp.write("\n")
                Current_Time = Log.Get_current_Time(self)  # Call a Function to Get Current Timings
                fp.write(str(Current_Time))
                fp.write("Expected PriceAfterIncrement ->> " + str(Expected_Price))
                fp.write("\n")
                Current_Time = Log.Get_current_Time(self)  # Call a Function to Get Current Timings
                fp.write(str(Current_Time))
                fp.write("Actual PriceAfterIncrement ->> " + str(Actual_Price))
                fp.write("\n")
                Current_Time = Log.Get_current_Time(self)  # Call a Function to Get Current Timings
                fp.write(str(Current_Time))
                fp.write("Result ->> " + str(Result))
                fp.write("\n")
                Current_Time = Log.Get_current_Time(self)  # Call a Function to Get Current Timings
                fp.write(str(Current_Time))
                fp.write("***************** - *******************")
            Log.Write_UP_ARROW_INCREMENT_CASES_Excel(self, Path, TC_NO, Test_Case, DgtsToselect, WINDOW_NAME, Result,row_value, Expected_Price, Actual_Price,Exch)

    # ----------------------Function to write log in "TC_PRICE_CURSR_DECREMENT_BUY and SELLORDENTRY.txt" --------------------#
    def Write_DOWN_ARROW_DECREMENT_CASES(self,Path, TC_NO,WINDOW_NAME,Test_Case,Exch,row_value, DgtsToselect, Result,Expected_Price, Actual_Price):
        if str(WINDOW_NAME) == "Buy Order Entry":
            file_name = "TC_PRICE_CURSR_DECREMENT_BUYORDENTRY.txt"                 # Initialize Name of the test file
            TestInfo_sheet = Get_File_Path("TESTCASE.xlsx", 1)  # Method to Active the TestData.xlsx sheet
            # print(str(TC_NO) + "----" + str(row_value))
            with (open(os.path.join(Path, file_name), 'a') as fp):
                fp.write("\n")
                Current_Time = Log.Get_current_Time(self)                          # Call a Function to Get Current Timings
                fp.write(str(Current_Time))
                fp.write("Window Name ->> " + str(WINDOW_NAME))
                fp.write("\n")
                Current_Time = Log.Get_current_Time(self)  # Call a Function to Get Current Timings
                fp.write(str(Current_Time))
                fp.write("Scenario ->> Price " + str(DgtsToselect) + " using Down Arrow ")
                fp.write("\n")
                Current_Time = Log.Get_current_Time(self)  # Call a Function to Get Current Timings
                fp.write(str(Current_Time))
                fp.write("Expected PriceAfterDecrement ->> " + str(Expected_Price))
                fp.write("\n")
                Current_Time = Log.Get_current_Time(self)  # Call a Function to Get Current Timings
                fp.write(str(Current_Time))
                fp.write("Actual PriceAfterDecrement ->> " + str(Actual_Price))
                fp.write("\n")
                Current_Time = Log.Get_current_Time(self)  # Call a Function to Get Current Timings
                fp.write(str(Current_Time))
                fp.write("Result ->> " + str(Result))
                fp.write("\n")
                Current_Time = Log.Get_current_Time(self)  # Call a Function to Get Current Timings
                fp.write(str(Current_Time))
                fp.write("***************** - *******************")
            Log.Write_DOWN_ARROW_DECREMENT_CASES_Excel(self, Path, TC_NO, Test_Case, DgtsToselect, WINDOW_NAME, Result,row_value, Expected_Price, Actual_Price,Exch)
        elif str(WINDOW_NAME) == "Sell Order Entry":
            file_name = "TC_PRICE_CURSR_DECREMENT_SELLORDENTRY.txt"                # Initialize Name of the test file
            TestInfo_sheet = Get_File_Path("TESTCASE.xlsx", 1)  # Method to Active the TestData.xlsx sheet
            # print(str(TC_NO) + "----" + str(row_value))
            with (open(os.path.join(Path, file_name), 'a') as fp):
                fp.write("\n")
                Current_Time = Log.Get_current_Time(self)                            # Call a Function to Get Current Timings
                fp.write(str(Current_Time))
                fp.write("Window Name ->> " + str(WINDOW_NAME))
                fp.write("\n")
                Current_Time = Log.Get_current_Time(self)  # Call a Function to Get Current Timings
                fp.write(str(Current_Time))
                fp.write("Scenario ->> Price " + str(DgtsToselect) + " using Down Arrow ")
                fp.write("\n")
                Current_Time = Log.Get_current_Time(self)  # Call a Function to Get Current Timings
                fp.write(str(Current_Time))
                fp.write("Expected PriceAfterDecrement ->> " + str(Expected_Price))
                fp.write("\n")
                Current_Time = Log.Get_current_Time(self)  # Call a Function to Get Current Timings
                fp.write(str(Current_Time))
                fp.write("Actual PriceAfterDecrement ->> " + str(Actual_Price))
                fp.write("\n")
                Current_Time = Log.Get_current_Time(self)  # Call a Function to Get Current Timings
                fp.write(str(Current_Time))
                fp.write("Result ->> " + str(Result))
                fp.write("\n")
                Current_Time = Log.Get_current_Time(self)  # Call a Function to Get Current Timings
                fp.write(str(Current_Time))
                fp.write("***************** - *******************")
            Log.Write_DOWN_ARROW_DECREMENT_CASES_Excel(self,Path, TC_NO, Test_Case, DgtsToselect, WINDOW_NAME, Result,row_value, Expected_Price, Actual_Price,Exch)

    # ----------------------Function to write log in "TC_PRICE_CURSR_RIGHTARROW_INSERT_BUY and SELLORDENTRY.txt" --------------------#
    def Write_RIGHTARROW_AND_INSERT_CASES(self,Path,TC_NO,WINDOW_NAME,TestCase,Input_Price,DgtsToselect,Position,Expected_Price,Actual_Price,Result,row_value):
        if str(WINDOW_NAME) == "Buy Order Entry":
            file_name = "TC_PRICE_CURSR_RIGHTARROW_INSERT_BUYORDENTRY.txt"  # Initialize Name of the test file
            TestInfo_sheet = Get_File_Path("TESTCASE.xlsx", 1)  # Method to Active the TestData.xlsx sheet
            #print(str(TC_NO) + "----" + str(row_value))
            with (open(os.path.join(Path, file_name), 'a') as fp):
                fp.write("\n")
                Current_Time = Log.Get_current_Time(self)           # Call a Function to Get Current Timings
                fp.write(str(Current_Time))
                fp.write("Window Name ->> "+ str(WINDOW_NAME))
                fp.write("\n")
                Current_Time = Log.Get_current_Time(self)                  # Call a Function to Get Current Timings
                fp.write(str(Current_Time))
                fp.write("Scenario ->> "+ str(DgtsToselect)+" by moving right arrow and Position of digit is "+str(Position))
                fp.write("\n")
                Current_Time = Log.Get_current_Time(self)                 # Call a Function to Get Current Timings
                fp.write(str(Current_Time))
                fp.write("Expected PriceAfterInsert ->> "+ str(Expected_Price))
                fp.write("\n")
                Current_Time = Log.Get_current_Time(self)                 # Call a Function to Get Current Timings
                fp.write(str(Current_Time))
                fp.write("Actual PriceAfterInsert ->> "+ str(Actual_Price))
                fp.write("\n")
                Current_Time = Log.Get_current_Time(self)                  # Call a Function to Get Current Timings
                fp.write(str(Current_Time))
                fp.write("Result ->> " + str(Result))
                fp.write("\n")
                Current_Time = Log.Get_current_Time(self)                  # Call a Function to Get Current Timings
                fp.write(str(Current_Time))
                fp.write("***************** - *******************")
            Log.Write_RIGHTARROW_AND_INSERT_CASES_Excel(self,Path, TC_NO, TestCase, DgtsToselect, WINDOW_NAME, Result,row_value, Input_Price, Position, Expected_Price, Actual_Price)
        elif str(WINDOW_NAME) == "Sell Order Entry":
            file_name = "TC_PRICE_CURSR_RIGHTARROW_INSERT_SELLORDENTRY.txt"  # Initialize Name of the test file
            TestInfo_sheet = Get_File_Path("TESTCASE.xlsx", 1)  # Method to Active the TestData.xlsx sheet
            # print(str(TC_NO) + "----" + str(row_value))
            with (open(os.path.join(Path, file_name), 'a') as fp):
                fp.write("\n")
                Current_Time = Log.Get_current_Time(self)  # Call a Function to Get Current Timings
                fp.write(str(Current_Time))
                fp.write("Window Name ->> " + str(WINDOW_NAME))
                fp.write("\n")
                Current_Time = Log.Get_current_Time(self)  # Call a Function to Get Current Timings
                fp.write(str(Current_Time))
                fp.write("Scenario ->> " + str(DgtsToselect) + "by moving right arrow and Position of digit is " + str(Position))
                fp.write("\n")
                Current_Time = Log.Get_current_Time(self)  # Call a Function to Get Current Timings
                fp.write(str(Current_Time))
                fp.write("Expected PriceAfterInsert ->> " + str(Expected_Price))
                fp.write("\n")
                Current_Time = Log.Get_current_Time(self)  # Call a Function to Get Current Timings
                fp.write(str(Current_Time))
                fp.write("Actual PriceAfterInsert ->> " + str(Actual_Price))
                fp.write("\n")
                Current_Time = Log.Get_current_Time(self)  # Call a Function to Get Current Timings
                fp.write(str(Current_Time))
                fp.write("Result ->> " + str(Result))
                fp.write("\n")
                Current_Time = Log.Get_current_Time(self)  # Call a Function to Get Current Timings
                fp.write(str(Current_Time))
                fp.write("***************** - *******************")
            Log.Write_RIGHTARROW_AND_INSERT_CASES_Excel(self,Path, TC_NO, TestCase, DgtsToselect, WINDOW_NAME, Result,row_value, Input_Price, Position, Expected_Price, Actual_Price)
    # -------------------------------------Function to write Excel of SELECT AND EDIT  cases----------------------------------------------#
    def Write_SELECT_AND_EDIT_CASES_Excel(self,Path, TC_NO, TestCase,DgtsToselect,WINDOW_NAME,Result,row_value,Input_Price,Position,Expected_Price,Actual_Price):                  #Method to update Testcase result status in "TestResults.xlsx"
        print("Row value in excel:" + str(row_value))
        Result_Excel_Sheet = os.path.join(Path, "TestResults.xlsx")                                              # "Result_Excel_Sheet" all testcases status will be stored
        target_workbook = openpyxl.load_workbook(Result_Excel_Sheet)
        if "SELECT_AND_EDIT" not in target_workbook.sheetnames:
            target_sheet = target_workbook.create_sheet(title="SELECT_AND_EDIT")  # This creates Sheet2 with a specific name
            target_sheet['A1'] = 'TESTCASE_NUMBER'
            target_sheet['B1'] = 'TEST_SCENARIO'
            target_sheet['C1'] = 'DigitsToSelect'
            target_sheet['D1'] = 'WINDOW_NAME'
            target_sheet['E1'] = 'INPUT'
            target_sheet['F1'] = 'POSITION'
            target_sheet['G1'] = 'EXPECTED RESULT'
            target_sheet['H1'] = 'ACTUAL RESULT'
            target_sheet['I1'] = 'RESULT'
        else:
            target_sheet = target_workbook["SELECT_AND_EDIT"]  # Access Sheet2
        next_row = target_sheet.max_row + 1                    #Search for next empty row and get row value
        # Paste the value into the target cell
        target_sheet.cell(row=next_row, column=1, value= str(TC_NO))
        target_sheet.cell(row=next_row, column=2, value= str(TestCase))
        target_sheet.cell(row=next_row, column=3, value= str(DgtsToselect))
        target_sheet.cell(row=next_row, column=4, value= str(WINDOW_NAME))
        target_sheet.cell(row=next_row, column=5, value=str(Input_Price))
        target_sheet.cell(row=next_row, column=6, value=str(Position))
        target_sheet.cell(row=next_row, column=7, value=str(Expected_Price))
        target_sheet.cell(row=next_row, column=8, value=str(Actual_Price))
        target_sheet.cell(row=next_row, column=9, value=str(Result))
        target_workbook.save(Result_Excel_Sheet)                                             # Save the target Excel file with the updated value

    # -------------------------------------Function to write Excel of DELETE_AND_EDIT  cases----------------------------------------------#
    def Write_DELETE_AND_EDIT_CASES_Excel(self, Path, TC_NO, Test_Case, DgtsToselect, Window_Name, Result,row_value, Input_Price, Position, Expected_Price, Actual_Price):                  #Method to update Testcase result status in "TestResults.xlsx"
        Result_Excel_Sheet = os.path.join(Path, "TestResults.xlsx")                                              # "Result_Excel_Sheet" all testcases status will be stored
        target_workbook = openpyxl.load_workbook(Result_Excel_Sheet)
        if "DELETE_AND_EDIT" not in target_workbook.sheetnames:
            target_sheet = target_workbook.create_sheet(title="DELETE_AND_EDIT")  # This creates Sheet2 with a specific name
            target_sheet['A1'] = 'TESTCASE_NUMBER'
            target_sheet['B1'] = 'TEST_SCENARIO'
            target_sheet['C1'] = 'DigitsToSelect'
            target_sheet['D1'] = 'WINDOW_NAME'
            target_sheet['E1'] = 'INPUT'
            target_sheet['F1'] = 'POSITION'
            target_sheet['G1'] = 'EXPECTED RESULT'
            target_sheet['H1'] = 'ACTUAL RESULT'
            target_sheet['I1'] = 'RESULT'
        else:
            target_sheet = target_workbook["DELETE_AND_EDIT"]  # Access Sheet2
        next_row = target_sheet.max_row + 1                    # Search for next empty row and get row value
        # Paste the value into the target cell
        target_sheet.cell(row=next_row, column=1, value=str(TC_NO))
        target_sheet.cell(row=next_row, column=2, value=str(Test_Case))
        target_sheet.cell(row=next_row, column=3, value=str(DgtsToselect))
        target_sheet.cell(row=next_row, column=4, value=str(Window_Name))
        target_sheet.cell(row=next_row, column=5, value=str(Input_Price))
        target_sheet.cell(row=next_row, column=6, value=str(Position))
        target_sheet.cell(row=next_row, column=7, value=str(Expected_Price))
        target_sheet.cell(row=next_row, column=8, value=str(Actual_Price))
        target_sheet.cell(row=next_row, column=9, value=str(Result))
        target_workbook.save(Result_Excel_Sheet)                                             # Save the target Excel file with the updated value

    # -------------------------------------Function to write Excel of BACKSPACE_AND_EDIT  cases----------------------------------------------#
    def Write_BACKSPACE_AND_EDIT_CASES_Excel(self, Path, TC_NO, Test_Case, DgtsToselect, Window_Name, Result,row_value, Input_Price, Position, Expected_Price, Actual_Price):                  #Method to update Testcase result status in "TestResults.xlsx"
        Result_Excel_Sheet = os.path.join(Path, "TestResults.xlsx")                                              # "Result_Excel_Sheet" all testcases status will be stored
        target_workbook = openpyxl.load_workbook(Result_Excel_Sheet)
        if "BACKSPACE_AND_EDIT" not in target_workbook.sheetnames:
            target_sheet = target_workbook.create_sheet(title="BACKSPACE_AND_EDIT")  # This creates Sheet2 with a specific name
            target_sheet['A1'] = 'TESTCASE_NUMBER'
            target_sheet['B1'] = 'TEST_SCENARIO'
            target_sheet['C1'] = 'DigitsToSelect'
            target_sheet['D1'] = 'WINDOW_NAME'
            target_sheet['E1'] = 'INPUT'
            target_sheet['F1'] = 'POSITION'
            target_sheet['G1'] = 'EXPECTED RESULT'
            target_sheet['H1'] = 'ACTUAL RESULT'
            target_sheet['I1'] = 'RESULT'
        else:
            target_sheet = target_workbook["BACKSPACE_AND_EDIT"]  # Access Sheet2
        next_row = target_sheet.max_row + 1                       # Search for next empty row and get row value
        # Paste the value into the target cell
        target_sheet.cell(row=next_row, column=1, value=str(TC_NO))
        target_sheet.cell(row=next_row, column=2, value=str(Test_Case))
        target_sheet.cell(row=next_row, column=3, value=str(DgtsToselect))
        target_sheet.cell(row=next_row, column=4, value=str(Window_Name))
        target_sheet.cell(row=next_row, column=5, value=str(Input_Price))
        target_sheet.cell(row=next_row, column=6, value=str(Position))
        target_sheet.cell(row=next_row, column=7, value=str(Expected_Price))
        target_sheet.cell(row=next_row, column=8, value=str(Actual_Price))
        target_sheet.cell(row=next_row, column=9, value=str(Result))
        target_workbook.save(Result_Excel_Sheet)
    # -------------------------------------Function to write Excel of UP_ARROW_INCREMENT  cases----------------------------------------------#
    def Write_UP_ARROW_INCREMENT_CASES_Excel(self,Path, TC_NO, Test_Case, DgtsToselect, Window_Name, Result,row_value,Expected_Price, Actual_Price,Exch):
        Result_Excel_Sheet = os.path.join(Path, "TestResults.xlsx")                 # "Result_Excel_Sheet" all testcases status will be stored
        target_workbook = openpyxl.load_workbook(Result_Excel_Sheet)
        if "UP_ARROW_INCREMENT" not in target_workbook.sheetnames:
            target_sheet = target_workbook.create_sheet(title="UP_ARROW_INCREMENT")  # This creates Sheet2 with a specific name
            target_sheet['A1'] = 'TESTCASE_NUMBER'
            target_sheet['B1'] = 'TEST_SCENARIO'
            target_sheet['C1'] = 'Exchange'
            target_sheet['D1'] = 'DigitsToSelect'
            target_sheet['E1'] = 'WINDOW_NAME'
            target_sheet['F1'] = 'EXPECTED RESULT'
            target_sheet['G1'] = 'ACTUAL RESULT'
            target_sheet['H1'] = 'RESULT'
        else:
            target_sheet = target_workbook["UP_ARROW_INCREMENT"]  # Access Sheet2
        next_row = target_sheet.max_row + 1                       # Search for next empty row and get row value
        # Paste the value into the target cell
        target_sheet.cell(row=next_row, column=1, value=str(TC_NO))
        target_sheet.cell(row=next_row, column=2, value=str(Test_Case))
        target_sheet.cell(row=next_row, column=3, value=str(Exch))
        target_sheet.cell(row=next_row, column=4, value=str(DgtsToselect))
        target_sheet.cell(row=next_row, column=5, value=str(Window_Name))
        target_sheet.cell(row=next_row, column=6, value=str(Expected_Price))
        target_sheet.cell(row=next_row, column=7, value=str(Actual_Price))
        target_sheet.cell(row=next_row, column=8, value=str(Result))
        target_workbook.save(Result_Excel_Sheet)

    # -------------------------------------Function to write Excel of DOWN_ARROW_INCREMENT  cases----------------------------------------------#
    def Write_DOWN_ARROW_DECREMENT_CASES_Excel(self, Path, TC_NO, Test_Case, DgtsToselect, Window_Name, Result,row_value, Expected_Price, Actual_Price,Exch):
        Result_Excel_Sheet = os.path.join(Path,"TestResults.xlsx")  # "Result_Excel_Sheet" all testcases status will be stored
        target_workbook = openpyxl.load_workbook(Result_Excel_Sheet)
        if "DOWN_ARROW_DECREMENT" not in target_workbook.sheetnames:
            target_sheet = target_workbook.create_sheet(title="DOWN_ARROW_DECREMENT")  # This creates Sheet2 with a specific name
            target_sheet['A1'] = 'TESTCASE_NUMBER'
            target_sheet['B1'] = 'TEST_SCENARIO'
            target_sheet['C1'] = 'Exchange'
            target_sheet['D1'] = 'DigitsToSelect'
            target_sheet['E1'] = 'WINDOW_NAME'
            target_sheet['F1'] = 'EXPECTED RESULT'
            target_sheet['G1'] = 'ACTUAL RESULT'
            target_sheet['H1'] = 'RESULT'
        else:
            target_sheet = target_workbook["DOWN_ARROW_DECREMENT"]  # Access Sheet2
        next_row = target_sheet.max_row + 1                         # Search for next empty row and get row value
        # Paste the value into the target cell
        target_sheet.cell(row=next_row, column=1, value=str(TC_NO))
        target_sheet.cell(row=next_row, column=2, value=str(Test_Case))
        target_sheet.cell(row=next_row, column=3, value=str(Exch))
        target_sheet.cell(row=next_row, column=4, value=str(DgtsToselect))
        target_sheet.cell(row=next_row, column=5, value=str(Window_Name))
        target_sheet.cell(row=next_row, column=6, value=str(Expected_Price))
        target_sheet.cell(row=next_row, column=7, value=str(Actual_Price))
        target_sheet.cell(row=next_row, column=8, value=str(Result))
        target_workbook.save(Result_Excel_Sheet)

    # -------------------------------------Function to write Excel of RIGHTARROW_AND_INSERT  cases---------------------------------------------
    def Write_RIGHTARROW_AND_INSERT_CASES_Excel(self, Path, TC_NO, Test_Case, DgtsToselect, Window_Name, Result,row_value, Input_Price, Position, Expected_Price, Actual_Price):
        Result_Excel_Sheet = os.path.join(Path,"TestResults.xlsx")  # "Result_Excel_Sheet" all testcases status will be stored
        target_workbook = openpyxl.load_workbook(Result_Excel_Sheet)
        # target_sheet = target_workbook["LMT_ORDER_TC_STATUS"]
        # row_value = str(TC_NO)
        if "RIGHT_ARROW_INSERT" not in target_workbook.sheetnames:
            target_sheet = target_workbook.create_sheet(title="RIGHT_ARROW_INSERT")  # This creates Sheet2 with a specific name
            target_sheet['A1'] = 'TESTCASE_NUMBER'
            target_sheet['B1'] = 'TEST_SCENARIO'
            target_sheet['C1'] = 'DigitsToSelect'
            target_sheet['D1'] = 'WINDOW_NAME'
            target_sheet['E1'] = 'INPUT'
            target_sheet['F1'] = 'POSITION'
            target_sheet['G1'] = 'EXPECTED RESULT'
            target_sheet['H1'] = 'ACTUAL RESULT'
            target_sheet['I1'] = 'RESULT'
        else:
            target_sheet = target_workbook["RIGHT_ARROW_INSERT"]  # Access Sheet2
        next_row = target_sheet.max_row + 1                       # Search for next empty row and get row value
        # Paste the value into the target cell
        target_sheet.cell(row=next_row, column=1, value=str(TC_NO))
        target_sheet.cell(row=next_row, column=2, value=str(Test_Case))
        target_sheet.cell(row=next_row, column=3, value=str(DgtsToselect))
        target_sheet.cell(row=next_row, column=4, value=str(Window_Name))
        target_sheet.cell(row=next_row, column=5, value=str(Input_Price))
        target_sheet.cell(row=next_row, column=6, value=str(Position))
        target_sheet.cell(row=next_row, column=7, value=str(Expected_Price))
        target_sheet.cell(row=next_row, column=8, value=str(Actual_Price))
        target_sheet.cell(row=next_row, column=9, value=str(Result))
        target_workbook.save(Result_Excel_Sheet)
#=============================================================================================================================================================#
    #-------------------------------------------Function to Get Current Time stamp to write a log ------------------------------------------------------------#
    def Get_current_Time(self):                                       #Function to Get Current Timing to print in a Result file
            current_second = datetime.now().second                        #Initialize current sec in a variable
            current_minute = datetime.now().minute                        #Initialize current min in a variable
            current_hour = datetime.now().hour                            #Initialize current hour in a variable
            current_day = datetime.now().day                              #Initialize current Date in a variable
            current_month = datetime.now().month                          #Initialize current Month in a variable
            current_year = datetime.now().year                            #Initialize current Year in a variable
            current_Time = str(current_day) + "/" + str(current_month) + "/" + str(current_year) + "  " + str(current_hour) + ":" + str(current_minute) + ":" + str(current_second) + " "
            return current_Time


